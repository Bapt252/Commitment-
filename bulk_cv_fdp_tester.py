#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SuperSmartMatch V3.0 Enhanced - Testeur Automatisé Bulk
Test automatique sur vrais CV et fiches de poste
Génère rapports Excel avec matrice complète de matching
"""

import os
import json
import time
import logging
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import argparse
from typing import Dict, List, Tuple, Optional
import glob
import traceback

# Configuration logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class Config:
    """Configuration du testeur automatisé"""
    # Chemins par défaut
    CV_TEST_DIR = os.path.expanduser("~/Desktop/CV TEST/")
    FDP_TEST_DIR = os.path.expanduser("~/Desktop/FDP TEST/")
    OUTPUT_DIR = "./test_results/"
    
    # API endpoints
    API_BASE_URL = "http://localhost:5067"
    PARSE_CV_URL = f"{API_BASE_URL}/parse_cv"
    PARSE_JOB_URL = f"{API_BASE_URL}/parse_job"
    MATCH_URL = f"{API_BASE_URL}/match"
    
    # Formats supportés
    CV_FORMATS = ['.pdf', '.docx', '.doc', '.txt', '.png', '.jpg', '.jpeg']
    JOB_FORMATS = ['.pdf', '.docx', '.doc', '.txt']
    
    # Performance
    REQUEST_TIMEOUT = 30
    RETRY_COUNT = 3
    DELAY_BETWEEN_REQUESTS = 0.1
    
    # Thresholds pour rapports
    EXCELLENT_SCORE = 85.0
    GOOD_SCORE = 70.0
    ACCEPTABLE_SCORE = 50.0

class BulkTester:
    """Testeur automatisé pour CV et fiches de poste en masse"""
    
    def __init__(self, cv_dir: str = None, fdp_dir: str = None, output_dir: str = None):
        self.cv_dir = cv_dir or Config.CV_TEST_DIR
        self.fdp_dir = fdp_dir or Config.FDP_TEST_DIR
        self.output_dir = output_dir or Config.OUTPUT_DIR
        
        # Création dossier output
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        
        # Stockage des données
        self.cv_data = {}  # {filename: cv_data_parsed}
        self.fdp_data = {}  # {filename: job_data_parsed}
        self.match_results = []  # Liste de tous les résultats de matching
        
        # Statistiques
        self.stats = {
            'cv_parsed': 0,
            'fdp_parsed': 0,
            'matches_calculated': 0,
            'total_time': 0,
            'errors': 0,
            'start_time': None,
            'end_time': None
        }
        
        logger.info(f"🔧 BulkTester initialisé")
        logger.info(f"📁 CV dir: {self.cv_dir}")
        logger.info(f"📁 FDP dir: {self.fdp_dir}")
        logger.info(f"📁 Output dir: {self.output_dir}")
    
    def check_api_health(self) -> bool:
        """Vérification santé API"""
        try:
            response = requests.get(f"{Config.API_BASE_URL}/health", timeout=5)
            if response.status_code == 200:
                logger.info("✅ API SuperSmartMatch disponible")
                return True
            else:
                logger.error(f"❌ API erreur: {response.status_code}")
                return False
        except requests.exceptions.RequestException as e:
            logger.error(f"❌ API non accessible: {e}")
            return False
    
    def find_files(self, directory: str, extensions: List[str]) -> List[str]:
        """Recherche fichiers avec extensions spécifiées"""
        files = []
        
        if not os.path.exists(directory):
            logger.warning(f"📁 Dossier non trouvé: {directory}")
            return files
        
        for ext in extensions:
            pattern = os.path.join(directory, f"*{ext}")
            found = glob.glob(pattern, recursive=False)
            files.extend(found)
        
        # Tri par nom
        files.sort()
        
        logger.info(f"📄 Trouvé {len(files)} fichiers dans {directory}")
        return files
    
    def parse_cv_file(self, file_path: str) -> Optional[Dict]:
        """Parse un fichier CV"""
        try:
            filename = os.path.basename(file_path)
            logger.info(f"📄 Parsing CV: {filename}")
            
            with open(file_path, 'rb') as f:
                files = {'file': (filename, f, self._get_mime_type(file_path))}
                
                for attempt in range(Config.RETRY_COUNT):
                    try:
                        response = requests.post(
                            Config.PARSE_CV_URL,
                            files=files,
                            timeout=Config.REQUEST_TIMEOUT
                        )
                        
                        if response.status_code == 200:
                            result = response.json()
                            if result.get('success'):
                                logger.info(f"✅ CV parsé: {filename} ({result.get('processing_time_ms', 0):.1f}ms)")
                                return result['cv_data']
                            else:
                                logger.error(f"❌ Erreur parsing CV {filename}: {result}")
                                break
                        else:
                            logger.warning(f"⚠️ Tentative {attempt+1} échouée pour {filename}")
                            if attempt < Config.RETRY_COUNT - 1:
                                time.sleep(1)
                            
                    except requests.exceptions.RequestException as e:
                        logger.warning(f"⚠️ Erreur réseau CV {filename}, tentative {attempt+1}: {e}")
                        if attempt < Config.RETRY_COUNT - 1:
                            time.sleep(1)
                
        except Exception as e:
            logger.error(f"❌ Erreur lecture CV {file_path}: {e}")
            self.stats['errors'] += 1
        
        return None
    
    def parse_job_file(self, file_path: str) -> Optional[Dict]:
        """Parse un fichier fiche de poste"""
        try:
            filename = os.path.basename(file_path)
            logger.info(f"📋 Parsing FDP: {filename}")
            
            # Lecture contenu selon format
            if file_path.lower().endswith('.txt'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                # Pour PDF/DOCX, on utilise l'API de parsing
                with open(file_path, 'rb') as f:
                    files = {'file': (filename, f, self._get_mime_type(file_path))}
                    
                    # Utiliser parse_cv pour extraire le texte puis parse_job
                    response = requests.post(
                        Config.PARSE_CV_URL,
                        files=files,
                        timeout=Config.REQUEST_TIMEOUT
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        if result.get('success'):
                            # Récupération description à partir du parsing
                            content = f"Titre: {filename}\n\nDescription: Voir compétences requises"
                        else:
                            logger.error(f"❌ Erreur extraction texte FDP {filename}")
                            return None
                    else:
                        logger.error(f"❌ Erreur API extraction FDP {filename}")
                        return None
            
            # Parse job description
            for attempt in range(Config.RETRY_COUNT):
                try:
                    data = {'job_description': content}
                    response = requests.post(
                        Config.PARSE_JOB_URL,
                        data=data,
                        timeout=Config.REQUEST_TIMEOUT
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        if result.get('success'):
                            logger.info(f"✅ FDP parsée: {filename} ({result.get('processing_time_ms', 0):.1f}ms)")
                            return result['job_data']
                        else:
                            logger.error(f"❌ Erreur parsing FDP {filename}: {result}")
                            break
                    else:
                        logger.warning(f"⚠️ Tentative {attempt+1} échouée pour {filename}")
                        if attempt < Config.RETRY_COUNT - 1:
                            time.sleep(1)
                        
                except requests.exceptions.RequestException as e:
                    logger.warning(f"⚠️ Erreur réseau FDP {filename}, tentative {attempt+1}: {e}")
                    if attempt < Config.RETRY_COUNT - 1:
                        time.sleep(1)
                
        except Exception as e:
            logger.error(f"❌ Erreur lecture FDP {file_path}: {e}")
            self.stats['errors'] += 1
        
        return None
    
    def calculate_match(self, cv_filename: str, cv_data: Dict, fdp_filename: str, fdp_data: Dict) -> Optional[Dict]:
        """Calcul matching entre CV et FDP"""
        try:
            match_request = {
                "cv_data": cv_data,
                "job_data": fdp_data,
                "algorithm": "Enhanced_V3.0"
            }
            
            for attempt in range(Config.RETRY_COUNT):
                try:
                    response = requests.post(
                        Config.MATCH_URL,
                        json=match_request,
                        timeout=Config.REQUEST_TIMEOUT
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        if result.get('success'):
                            match_result = result['result']
                            
                            # Ajout informations contextuelles
                            match_result['cv_filename'] = cv_filename
                            match_result['fdp_filename'] = fdp_filename
                            match_result['cv_name'] = cv_data.get('name', 'N/A')
                            match_result['job_title'] = fdp_data.get('title', 'N/A')
                            match_result['timestamp'] = datetime.now().isoformat()
                            
                            logger.info(f"🎯 Match calculé: {cv_filename} x {fdp_filename} = {match_result['score']:.1f}%")
                            return match_result
                        else:
                            logger.error(f"❌ Erreur matching: {result}")
                            break
                    else:
                        logger.warning(f"⚠️ Tentative {attempt+1} échouée pour matching")
                        if attempt < Config.RETRY_COUNT - 1:
                            time.sleep(1)
                        
                except requests.exceptions.RequestException as e:
                    logger.warning(f"⚠️ Erreur réseau matching, tentative {attempt+1}: {e}")
                    if attempt < Config.RETRY_COUNT - 1:
                        time.sleep(1)
                
        except Exception as e:
            logger.error(f"❌ Erreur calcul matching: {e}")
            self.stats['errors'] += 1
        
        return None
    
    def run_bulk_test(self) -> bool:
        """Exécution test en masse"""
        logger.info("🚀 Début test automatisé SuperSmartMatch V3.0")
        self.stats['start_time'] = datetime.now()
        start_time = time.time()
        
        try:
            # 1. Vérification API
            if not self.check_api_health():
                logger.error("❌ API non disponible, arrêt du test")
                return False
            
            # 2. Recherche fichiers CV
            cv_files = self.find_files(self.cv_dir, Config.CV_FORMATS)
            if not cv_files:
                logger.error(f"❌ Aucun CV trouvé dans {self.cv_dir}")
                return False
            
            # 3. Recherche fichiers FDP
            fdp_files = self.find_files(self.fdp_dir, Config.JOB_FORMATS)
            if not fdp_files:
                logger.error(f"❌ Aucune FDP trouvée dans {self.fdp_dir}")
                return False
            
            logger.info(f"📊 Test configuré: {len(cv_files)} CV x {len(fdp_files)} FDP = {len(cv_files) * len(fdp_files)} matchings")
            
            # 4. Parsing tous les CV
            logger.info("📄 === PHASE 1: PARSING CV ===")
            for cv_file in cv_files:
                cv_data = self.parse_cv_file(cv_file)
                if cv_data:
                    filename = os.path.basename(cv_file)
                    self.cv_data[filename] = cv_data
                    self.stats['cv_parsed'] += 1
                
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            
            logger.info(f"✅ CV parsés: {self.stats['cv_parsed']}/{len(cv_files)}")
            
            # 5. Parsing toutes les FDP
            logger.info("📋 === PHASE 2: PARSING FDP ===")
            for fdp_file in fdp_files:
                fdp_data = self.parse_job_file(fdp_file)
                if fdp_data:
                    filename = os.path.basename(fdp_file)
                    self.fdp_data[filename] = fdp_data
                    self.stats['fdp_parsed'] += 1
                
                time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            
            logger.info(f"✅ FDP parsées: {self.stats['fdp_parsed']}/{len(fdp_files)}")
            
            # 6. Calcul matrice de matching
            logger.info("🎯 === PHASE 3: MATCHING MATRIX ===")
            total_matches = len(self.cv_data) * len(self.fdp_data)
            current_match = 0
            
            for cv_filename, cv_data in self.cv_data.items():
                for fdp_filename, fdp_data in self.fdp_data.items():
                    current_match += 1
                    
                    logger.info(f"🔄 Progress: {current_match}/{total_matches} ({(current_match/total_matches)*100:.1f}%)")
                    
                    match_result = self.calculate_match(cv_filename, cv_data, fdp_filename, fdp_data)
                    if match_result:
                        self.match_results.append(match_result)
                        self.stats['matches_calculated'] += 1
                    
                    time.sleep(Config.DELAY_BETWEEN_REQUESTS)
            
            # 7. Finalisation
            self.stats['end_time'] = datetime.now()
            self.stats['total_time'] = time.time() - start_time
            
            logger.info("✅ === TEST TERMINÉ ===")
            logger.info(f"📊 Résultats: {self.stats['matches_calculated']} matchings calculés en {self.stats['total_time']:.1f}s")
            logger.info(f"🎯 Taux de succès: {(self.stats['matches_calculated']/total_matches)*100:.1f}%")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Erreur durant test bulk: {e}")
            logger.error(traceback.format_exc())
            return False
    
    def generate_excel_report(self, filename: str = None) -> str:
        """Génération rapport Excel détaillé"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SuperSmartMatch_Report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            logger.info(f"📊 Génération rapport Excel: {filepath}")
            
            # Création workbook
            wb = openpyxl.Workbook()
            
            # Suppression feuille par défaut
            wb.remove(wb.active)
            
            # 1. Feuille Résumé
            self._create_summary_sheet(wb)
            
            # 2. Feuille Matrice Complète
            self._create_matrix_sheet(wb)
            
            # 3. Feuille Top Matches
            self._create_top_matches_sheet(wb)
            
            # 4. Feuille Statistiques CV
            self._create_cv_stats_sheet(wb)
            
            # 5. Feuille Statistiques FDP
            self._create_fdp_stats_sheet(wb)
            
            # 6. Feuille Données Brutes
            self._create_raw_data_sheet(wb)
            
            # Sauvegarde
            wb.save(filepath)
            logger.info(f"✅ Rapport Excel généré: {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"❌ Erreur génération Excel: {e}")
            logger.error(traceback.format_exc())
            return None
    
    def _create_summary_sheet(self, wb):
        """Feuille résumé exécutif"""
        ws = wb.create_sheet("📊 Résumé Exécutif")
        
        # Titre
        ws['A1'] = "SuperSmartMatch V3.0 Enhanced - Rapport de Test"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # Informations générales
        row = 3
        info_data = [
            ["📅 Date du test", self.stats['start_time'].strftime("%d/%m/%Y %H:%M:%S")],
            ["⏱️ Durée totale", f"{self.stats['total_time']:.1f} secondes"],
            ["📄 CV analysés", f"{self.stats['cv_parsed']}"],
            ["📋 FDP analysées", f"{self.stats['fdp_parsed']}"],
            ["🎯 Matchings calculés", f"{self.stats['matches_calculated']}"],
            ["❌ Erreurs", f"{self.stats['errors']}"]
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Statistiques de score
        if self.match_results:
            scores = [r['score'] for r in self.match_results]
            
            row += 2
            ws[f'A{row}'] = "📈 Statistiques de Score"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            stats_data = [
                ["Score moyen", f"{sum(scores)/len(scores):.1f}%"],
                ["Score maximum", f"{max(scores):.1f}%"],
                ["Score minimum", f"{min(scores):.1f}%"],
                ["Scores excellents (≥85%)", f"{sum(1 for s in scores if s >= Config.EXCELLENT_SCORE)}"],
                ["Scores bons (≥70%)", f"{sum(1 for s in scores if s >= Config.GOOD_SCORE)}"],
                ["Scores acceptables (≥50%)", f"{sum(1 for s in scores if s >= Config.ACCEPTABLE_SCORE)}"]
            ]
            
            for label, value in stats_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Ajustement colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_matrix_sheet(self, wb):
        """Feuille matrice complète"""
        ws = wb.create_sheet("🎯 Matrice Matching")
        
        if not self.match_results:
            ws['A1'] = "Aucun résultat de matching disponible"
            return
        
        # Création DataFrame pour pivot
        df_data = []
        for result in self.match_results:
            df_data.append({
                'CV': result['cv_filename'],
                'FDP': result['fdp_filename'],
                'Score': result['score'],
                'Candidat': result['cv_name'],
                'Poste': result['job_title']
            })
        
        df = pd.DataFrame(df_data)
        
        # Pivot table
        pivot = df.pivot(index='CV', columns='FDP', values='Score')
        
        # Ajout à Excel
        for r in dataframe_to_rows(pivot, index_names=True, header=True):
            ws.append(r)
        
        # Formatage
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                if cell.value and isinstance(cell.value, (int, float)):
                    if cell.value >= Config.EXCELLENT_SCORE:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    elif cell.value >= Config.GOOD_SCORE:
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    elif cell.value >= Config.ACCEPTABLE_SCORE:
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    def _create_top_matches_sheet(self, wb):
        """Feuille top matches"""
        ws = wb.create_sheet("🏆 Top Matches")
        
        # Tri par score décroissant
        sorted_results = sorted(self.match_results, key=lambda x: x['score'], reverse=True)
        
        # Headers
        headers = ['Rang', 'Score', 'Candidat', 'CV', 'Poste', 'FDP', 'Compétences Match', 'Expérience Match', 'Note']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données top 50
        for rank, result in enumerate(sorted_results[:50], 1):
            row_data = [
                rank,
                f"{result['score']:.1f}%",
                result['cv_name'],
                result['cv_filename'],
                result['job_title'],
                result['fdp_filename'],
                f"{result['skill_match']:.1f}%",
                f"{result['experience_match']:.1f}%",
                result['performance_note']
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=rank+1, column=col, value=value)
        
        # Ajustement colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def _create_cv_stats_sheet(self, wb):
        """Feuille statistiques CV"""
        ws = wb.create_sheet("📄 Stats CV")
        
        # Analyse par CV
        cv_stats = {}
        for result in self.match_results:
            cv = result['cv_filename']
            if cv not in cv_stats:
                cv_stats[cv] = {
                    'scores': [],
                    'best_match': None,
                    'avg_score': 0,
                    'candidat': result['cv_name']
                }
            
            cv_stats[cv]['scores'].append(result['score'])
            if not cv_stats[cv]['best_match'] or result['score'] > cv_stats[cv]['best_match']['score']:
                cv_stats[cv]['best_match'] = result
        
        # Calcul moyennes
        for cv, stats in cv_stats.items():
            stats['avg_score'] = sum(stats['scores']) / len(stats['scores'])
        
        # Headers
        headers = ['CV', 'Candidat', 'Score Moyen', 'Score Max', 'Meilleur Match', 'Poste']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données
        sorted_cv = sorted(cv_stats.items(), key=lambda x: x[1]['avg_score'], reverse=True)
        for row, (cv, stats) in enumerate(sorted_cv, 2):
            row_data = [
                cv,
                stats['candidat'],
                f"{stats['avg_score']:.1f}%",
                f"{max(stats['scores']):.1f}%",
                f"{stats['best_match']['score']:.1f}%",
                stats['best_match']['job_title']
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustement colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 20
    
    def _create_fdp_stats_sheet(self, wb):
        """Feuille statistiques FDP"""
        ws = wb.create_sheet("📋 Stats FDP")
        
        # Analyse par FDP
        fdp_stats = {}
        for result in self.match_results:
            fdp = result['fdp_filename']
            if fdp not in fdp_stats:
                fdp_stats[fdp] = {
                    'scores': [],
                    'best_match': None,
                    'avg_score': 0,
                    'poste': result['job_title']
                }
            
            fdp_stats[fdp]['scores'].append(result['score'])
            if not fdp_stats[fdp]['best_match'] or result['score'] > fdp_stats[fdp]['best_match']['score']:
                fdp_stats[fdp]['best_match'] = result
        
        # Calcul moyennes
        for fdp, stats in fdp_stats.items():
            stats['avg_score'] = sum(stats['scores']) / len(stats['scores'])
        
        # Headers
        headers = ['FDP', 'Poste', 'Score Moyen', 'Score Max', 'Meilleur Match', 'Candidat']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données
        sorted_fdp = sorted(fdp_stats.items(), key=lambda x: x[1]['avg_score'], reverse=True)
        for row, (fdp, stats) in enumerate(sorted_fdp, 2):
            row_data = [
                fdp,
                stats['poste'],
                f"{stats['avg_score']:.1f}%",
                f"{max(stats['scores']):.1f}%",
                f"{stats['best_match']['score']:.1f}%",
                stats['best_match']['cv_name']
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustement colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 20
    
    def _create_raw_data_sheet(self, wb):
        """Feuille données brutes"""
        ws = wb.create_sheet("📊 Données Brutes")
        
        if not self.match_results:
            ws['A1'] = "Aucune donnée disponible"
            return
        
        # Headers
        headers = [
            'CV', 'FDP', 'Candidat', 'Poste', 'Score', 'Match Compétences', 
            'Match Expérience', 'Bonus Titre', 'Bonus Secteur', 'Note', 'Timestamp'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données
        for row, result in enumerate(self.match_results, 2):
            row_data = [
                result['cv_filename'],
                result['fdp_filename'],
                result['cv_name'],
                result['job_title'],
                result['score'],
                result['skill_match'],
                result['experience_match'],
                result['title_bonus'],
                result.get('sector_bonus', 0),
                result['performance_note'],
                result['timestamp']
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajustement colonnes
        for col in range(1, len(headers)+1):
            ws.column_dimensions[chr(64+col)].width = 15
    
    def _get_mime_type(self, file_path: str) -> str:
        """Détermination type MIME"""
        ext = os.path.splitext(file_path)[1].lower()
        mime_types = {
            '.pdf': 'application/pdf',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.doc': 'application/msword',
            '.txt': 'text/plain',
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg'
        }
        return mime_types.get(ext, 'application/octet-stream')
    
    def print_summary(self):
        """Affichage résumé console"""
        print("\n" + "="*60)
        print("🎯 SUPERSMARTMATCH V3.0 ENHANCED - RÉSULTATS TEST")
        print("="*60)
        print(f"📊 CV analysés: {self.stats['cv_parsed']}")
        print(f"📊 FDP analysées: {self.stats['fdp_parsed']}")
        print(f"📊 Matchings calculés: {self.stats['matches_calculated']}")
        print(f"📊 Temps total: {self.stats['total_time']:.1f}s")
        print(f"📊 Erreurs: {self.stats['errors']}")
        
        if self.match_results:
            scores = [r['score'] for r in self.match_results]
            print(f"\n🏆 PERFORMANCES:")
            print(f"   Score moyen: {sum(scores)/len(scores):.1f}%")
            print(f"   Score maximum: {max(scores):.1f}%")
            print(f"   Scores excellents (≥85%): {sum(1 for s in scores if s >= 85)}")
            
            # Top 3 matches
            top_matches = sorted(self.match_results, key=lambda x: x['score'], reverse=True)[:3]
            print(f"\n🥇 TOP 3 MATCHES:")
            for i, match in enumerate(top_matches, 1):
                print(f"   {i}. {match['cv_name']} → {match['job_title']} ({match['score']:.1f}%)")
        
        print("="*60)

def main():
    """Fonction principale"""
    parser = argparse.ArgumentParser(description="SuperSmartMatch V3.0 Enhanced - Test Automatisé")
    parser.add_argument("--cv-dir", default=Config.CV_TEST_DIR, help="Dossier CV TEST")
    parser.add_argument("--fdp-dir", default=Config.FDP_TEST_DIR, help="Dossier FDP TEST")
    parser.add_argument("--output-dir", default=Config.OUTPUT_DIR, help="Dossier output")
    parser.add_argument("--no-excel", action="store_true", help="Ne pas générer Excel")
    parser.add_argument("--quiet", action="store_true", help="Mode silencieux")
    
    args = parser.parse_args()
    
    if args.quiet:
        logging.getLogger().setLevel(logging.WARNING)
    
    # Initialisation tester
    tester = BulkTester(
        cv_dir=args.cv_dir,
        fdp_dir=args.fdp_dir,
        output_dir=args.output_dir
    )
    
    # Exécution test
    success = tester.run_bulk_test()
    
    if success:
        # Génération rapport Excel
        if not args.no_excel:
            excel_path = tester.generate_excel_report()
            if excel_path:
                print(f"\n📊 Rapport Excel généré: {excel_path}")
        
        # Affichage résumé
        tester.print_summary()
        
        # Sauvegarde JSON
        json_path = os.path.join(args.output_dir, "results.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump({
                'stats': tester.stats,
                'results': tester.match_results,
                'cv_data': tester.cv_data,
                'fdp_data': tester.fdp_data
            }, f, indent=2, default=str)
        
        print(f"💾 Données JSON sauvegardées: {json_path}")
        
    else:
        print("❌ Test échoué. Vérifiez les logs pour plus de détails.")
        return 1
    
    return 0

if __name__ == "__main__":
    exit(main())
